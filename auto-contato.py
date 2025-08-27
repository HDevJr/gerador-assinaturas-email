import os
from PIL import Image, ImageDraw
from openpyxl import load_workbook
from ttf_opensans import OPENSANS_BOLD, opensans

def gerar_todos_cartoes(template_path, excel_path, output_dir):
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    font_name = OPENSANS_BOLD.imagefont(size=28)
    font_cargo = OPENSANS_BOLD.imagefont(size=18)
    font_tel_bold = OPENSANS_BOLD.imagefont(size=16)
    font_tel_regular = opensans(font_weight=400).imagefont(size=16)

    os.makedirs(output_dir, exist_ok=True)

    def draw_telefone(draw, txt, base_pos):
        if not txt or not txt.strip().startswith("("):
            draw.text(base_pos, txt, font=font_tel_bold, fill="#2f2f2f")
            return
        parts = txt.split(")")
        ddd = parts[0] + ")"
        numero = parts[1].strip() if len(parts) > 1 else ""
        w_ddd = draw.textbbox((0, 0), ddd, font=font_tel_regular)[2]
        draw.text(base_pos, ddd, font=font_tel_regular, fill="#2f2f2f")
        draw.text((base_pos[0] + w_ddd + 2, base_pos[1]), numero, font=font_tel_bold, fill="#2f2f2f")

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        nome, cargo, tel1, tel2 = row[:4]
        if not nome:
            print(f"Linha {idx}: nome vazio — pulando.")
            continue

        img = Image.open(template_path).convert("RGBA")
        draw = ImageDraw.Draw(img)

        pos_nome = (191, 81)
        pos_cargo = (193, 117)
        pos_tel1 = (198, 153)
        pos_tel2 = (195, 173)

        draw.text(pos_nome, str(nome), font=font_name, fill="#373737")

        # DEFINE A COR DO CAMPO CARGO
        # draw.text(pos_cargo, str(cargo or ""), font=font_cargo, fill="#25418c")
        draw.text(pos_cargo, str(cargo or ""), font=font_cargo, fill="#004039")
        # draw.text(pos_cargo, str(cargo or ""), font=font_cargo, fill="#941629")

        if tel1:
            draw_telefone(draw, str(tel1), pos_tel1)
        if tel2:
            draw_telefone(draw, str(tel2), pos_tel2)

        parts = str(nome).strip().split()
        if len(parts) >= 2:
            filename = f"{parts[0]}-{parts[-1]}.png"
        else:
            filename = f"{parts[0]}.png"
        # Remove caracteres inválidos do nome do arquivo
        safe_filename = "".join(c for c in filename if c.isalnum() or c in "._-")
        out_path = os.path.join(output_dir, safe_filename)

        img.save(out_path)
        print(f"Gerado (linha {idx}): {out_path}")

if __name__ == "__main__":
    gerar_todos_cartoes(
        template_path="template-contato.png",
        excel_path="usuarios.xlsx",
        output_dir="contato-usuarios"
    )
